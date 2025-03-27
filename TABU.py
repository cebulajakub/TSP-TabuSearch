import os
import random
import time
import numpy as np
import networkx as nx
import cv2
import keyboard
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import psutil
from openpyxl import Workbook

def calculate_cost(G, tour):
    total_cost = 0
    for i in range(0, len(tour) - 1):
        total_cost += G[tour[i]][tour[i + 1]]['weight']
    total_cost += G[tour[-1]][tour[0]]['weight']
    return total_cost

def generate_3opt_neighborhood(tour):
    neighborhood = []
    n = len(tour)
    for i in range(n):
        for j in range(i + 1, n):
            for k in range(j + 1, n):
                new_tour_1 = tour[:i] + tour[i:j][::-1] + tour[j:k][::-1] + tour[k:]
                new_tour_2 = tour[:i] + tour[j:k] + tour[i:j] + tour[k:]
                new_tour_3 = tour[:i] + tour[j:k][::-1] + tour[i:j] + tour[k:]
                neighborhood.extend([new_tour_1, new_tour_2, new_tour_3])
    return neighborhood

def generate_2opt_neighborhood(tour):
    neighborhood = []
    n = len(tour)
    for i in range(n - 1):
        for j in range(i + 2, n):
            new_tour = tour[:i] + tour[i:j][::-1] + tour[j:]
            neighborhood.append(new_tour)
    return neighborhood

def random_edge_swap(tour):
    n = len(tour)
    i, j = random.sample(range(n), 2)
    tour[i], tour[j] = tour[j], tour[i]
    return tour

def diversify_solution(tour, diversification_factor):
    n = len(tour)
    num_swaps = int(diversification_factor * n)
    for _ in range(num_swaps):
        i, j = random.sample(range(n), 2)
        tour[i], tour[j] = tour[j], tour[i]
    return tour

def greedy_solution(G):
    start_node = random.choice(list(G.nodes))
    unvisited = set(G.nodes)
    unvisited.remove(start_node)
    tour = [start_node]
    while unvisited:
        last_node = tour[-1]
        next_node = min(unvisited, key=lambda node: G[last_node][node]['weight'])
        tour.append(next_node)
        unvisited.remove(next_node)
    return tour

def dynamic_tabu_tenure(iteration, max_iterations, base_tenure=10):
    return base_tenure + int((iteration / max_iterations) * base_tenure * base_tenure)

def memory_usage():
    process = psutil.Process(os.getpid())
    return process.memory_info().rss

def tabu_search(G, max_iterations=1000, base_tabu_tenure=10, diversification_factor=0.15):
    nodes = list(G.nodes)
    if not nodes:
        raise ValueError("The graph is empty, no nodes to form a solution.")
    current_solution = greedy_solution(G)
    best_solution = current_solution
    best_cost = calculate_cost(G, best_solution)
    global_best_cost = best_cost
    global_best_solution = best_solution
    tabu_list = []
    iteration = 0
    no_improve_count = 0
    max_no_improve = 25
    history = set()
    tabu_tenure = 15
    while iteration < max_iterations:
        if no_improve_count > max_no_improve // 2:
            tabu_tenure = min(tabu_tenure + 5, 40)
        else:
            tabu_tenure = max(15, tabu_tenure - 2)
        if iteration % 10 == 0 or iteration % 11 == 0:
            neighborhood = generate_3opt_neighborhood(current_solution)
        else:
            neighborhood = generate_2opt_neighborhood(current_solution)
        best_move = None
        best_move_cost = float('inf')
        best_move_edges = None
        for move in neighborhood:
            move_edges = set(zip(move, move[1:] + [move[0]]))
            move_cost = calculate_cost(G, move)
            if tuple(move) not in history:
                if move_edges not in tabu_list or move_cost < best_cost:
                    if move_cost < best_move_cost:
                        best_move_cost = move_cost
                        best_move = move
                        best_move_edges = move_edges
        if best_move is not None:
            current_solution = best_move
            current_cost = best_move_cost
            if current_cost < best_cost:
                best_solution = current_solution
                best_cost = current_cost
                no_improve_count = 0
            else:
                no_improve_count += 1
            tabu_list.append(best_move_edges)
            if len(tabu_list) > tabu_tenure:
                tabu_list.pop(0)
                tabu_list.pop(0)
            history.add(tuple(current_solution))
            if no_improve_count > max_no_improve:
                current_solution = diversify_solution(current_solution, diversification_factor)
                current_cost = calculate_cost(G, current_solution)
                no_improve_count = 0
                best_cost = current_cost
            if current_cost < global_best_cost:
                global_best_cost = current_cost
                global_best_solution = current_solution
        iteration += 1
        if global_best_cost <= 7544.366 or iteration == max_iterations:
            save_to_excel(iteration, time.time() - start_time, global_best_cost)
            break
        if keyboard.is_pressed('q'):
            break
    return global_best_solution, global_best_cost

def load(path):
    tree = ET.parse(path)
    root = tree.getroot()
    G = nx.DiGraph()
    vertex_id = 0
    for vertex in root.findall('.//graph//vertex'):
        for edge in vertex.findall('edge'):
            edge_cost = float(edge.attrib.get('cost', 0))
            to_vertex = edge.text.strip()
            G.add_edge(vertex_id, int(to_vertex), weight=edge_cost)
        vertex_id += 1
    return G

def save_to_excel(iteration, time_elapsed, best_cost):
    try:
        wb = load_workbook('results.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Iteration', 'Time Elapsed (s)', 'Best Cost'])
    first_empty_row = sheet.max_row + 1
    sheet.cell(row=first_empty_row, column=1).value = iteration
    sheet.cell(row=first_empty_row, column=2).value = time_elapsed
    sheet.cell(row=first_empty_row, column=3).value = best_cost
    wb.save('results.xlsx')

if __name__ == "__main__":
    path = "berlin52.xml"
    G = load(path)
    try:
        for i in range(10000):
            start_time = time.time()
            best_solution, best_cost = tabu_search(G, max_iterations=1000, diversification_factor=0.15)
            end_time = time.time()
    except ValueError as e:
        print(e)
