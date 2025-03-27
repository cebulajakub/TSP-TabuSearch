import os
import random
import time
import numpy as np
import networkx as nx
import cv2
import keyboard
import xml.etree.ElementTree as ET

import openpyxl
import pandas as pd
from openpyxl import load_workbook


def calculate_cost(G, tour):
    total_cost = 0
    # print(f'Number of nodes { len(tour)} ')
    for i in range(0, len(tour) - 1):
        total_cost += G[tour[i]][tour[i + 1]]['weight']

    # last to first
    total_cost += G[tour[-1]][tour[0]]['weight']

    # print(f'Total cost is {total_cost}')

    return total_cost


'''
musimy miec permutacje:
normlanie bedzie ich  n!
ale nie musimy przegladac wszystkich 
tylko (n-1)!/2 zeby nie powtarzac
wybieramy randomowa permutacje  i szukamy najblizszych sasiadów (roziwązań) tzw. downhill
 najprosztsza metod 2-opt (zamiana 2 krawedzi liczymyu zamiane z całego zbioru a nie tylko z 2 losowych wierzchołkó) :

 Aby zaminiec trzeba zrobic liste najlepszych ruchów czyli policzyc wszystkie mozliwe trasy
 my mamy graf pelny wiec duzo liczenia to zajmie(chyba) i uwzgledniac tabu liste ( nwm jak ja narazie reprezentowac)


 3-opt złożoność n(n−1)(n−2)/6 ----- O(n^3)

 następnie wykonac zamiane policzyc koszt z uwzglednieniem tabu listy

 Wyznaczamy zbiory liczb któr mozna przestrawiac i dla kazdej pary wyznaczamy permutacje

 Algorytm unika oscylacji wokół optimum lokalnego dzięki przechowywaniu informacji o sprawdzonych już rozwiązaniach w postaci listy tabu (TL).


'''


# def generate_tabu_matrix(G):


def generate_3opt_neighborhood(tour):
    neighborhood = []
    n = len(tour)
    for i in range(n):
        for j in range(i + 1, n):
            for k in range(j + 1, n):
                new_tour_1 = tour[:i] + tour[i:j][::-1] + tour[j:k][::-1] + tour[k:]
                new_tour_2 = tour[:i] + tour[j:k] + tour[i:j] + tour[k:]
                new_tour_3 = tour[:i] + tour[j:k][::-1] + tour[i:j] + tour[k:]
                neighborhood.extend([new_tour_1, new_tour_2, new_tour_3])
    return neighborhood


def generate_2opt_neighborhood(tour):
    neighborhood = []
    n = len(tour)
    for i in range(n - 1):
        for j in range(i + 2, n):
            new_tour = tour[:i] + tour[i:j][::-1] + tour[j:]
            neighborhood.append(new_tour)
    return neighborhood


def random_edge_swap(tour):
    """Randomly swap two edges in the tour"""
    n = len(tour)
    i, j = random.sample(range(n), 2)  # Randomly select two different indices
    # Swap positions of the nodes in the tour
    tour[i], tour[j] = tour[j], tour[i]
    return tour



def diversify_solution(tour, diversification_factor):
    """Dywersyfikacja: losowo zmienia część trasy."""
    n = len(tour)
    num_swaps = int(diversification_factor * n)
    #print("Number of vertex swap :", num_swaps)
    for _ in range(num_swaps):
        i, j = random.sample(range(n), 2)
        tour[i], tour[j] = tour[j], tour[i]
    return tour




def greedy_solution(G):
    start_node = random.choice(list(G.nodes))
    unvisited = set(G.nodes)
    unvisited.remove(start_node)
    tour = [start_node]

    while unvisited:
        last_node = tour[-1]
        next_node = min(unvisited, key=lambda node: G[last_node][node]['weight'])
        tour.append(next_node)
        unvisited.remove(next_node)

    return tour

def dynamic_tabu_tenure(iteration, max_iterations, base_tenure=10):
    """Dynamiczne dostosowanie długości listy tabu."""
    return base_tenure + int((iteration / max_iterations) * base_tenure * base_tenure)

import psutil
def memory_usage():
    process = psutil.Process(os.getpid())
    return process.memory_info().rss  # Wartość w bajtach



def tabu_search(G, max_iterations=1000, base_tabu_tenure=10, diversification_factor=0.15):
    """Algorytm Tabu Search z adaptacyjną dywersyfikacją i dynamicznymi parametrami."""
    nodes = list(G.nodes)
    #print("nodes", nodes)
    if not nodes:
        raise ValueError("The graph is empty, no nodes to form a solution.")

    # Inicjalizacja
    current_solution = greedy_solution(G)
    best_solution = current_solution
    best_cost = calculate_cost(G, best_solution)
    global_best_cost = best_cost
    global_best_solution = best_solution
    tabu_list = []
    iteration = 0
    no_improve_count = 0
    max_no_improve = 25
    history = set()
    # last_diversification_iteration = -float('inf')  # Ostatnia iteracja z dywersyfikacją
    tabu_tenure = 15
    #print("Tabu Search started...\n")
    #print(f"Initial solution: {current_solution}, Initial cost: {best_cost}\n")

    while iteration < max_iterations:
        # Dynamiczne dostosowanie parametrow
        #tabu_tenure = max(15, min(dynamic_tabu_tenure(iteration, max_iterations), 50))
        if no_improve_count > max_no_improve // 2:
           tabu_tenure = min(tabu_tenure + 5, 40)
        else:
           tabu_tenure = max(15, tabu_tenure - 2)

        # max_no_improve = dynamic_max_no_improve(iteration, max_iterations)

        # Generowanie sasiedztwa
        if iteration % 10 == 0 or iteration % 11 == 0 :
            neighborhood = generate_3opt_neighborhood(current_solution)
        else:
            neighborhood = generate_2opt_neighborhood(current_solution)
        best_move = None
        best_move_cost = float('inf')
        best_move_edges = None

        # Przeszukiwanie sasiedztwa
        for move in neighborhood:
            move_edges = set(zip(move, move[1:] + [move[0]]))  # Zapisujemy krawedzie
            move_cost = calculate_cost(G, move)

            if tuple(move) not in history:
                # aspiracja
                if move_edges not in tabu_list or move_cost < best_cost:
                    if move_cost < best_move_cost:
                        best_move_cost = move_cost
                        best_move = move
                        best_move_edges = move_edges

        # Jeśli znaleziono ruch
        if best_move is not None:
            current_solution = best_move
            current_cost = best_move_cost

            # Aktualizacja najlepszego rozwiązania  gdy lepsze
            if current_cost < best_cost:
                best_solution = current_solution
                best_cost = current_cost
                no_improve_count = 0  # Reset licznika
            else:
                no_improve_count += 1
            #draw_graph_cv2(G, positions, current_solution)
            draw_graph_cv2(G, positions, global_best_solution)
            if iteration == 1:

                cv2.waitKey(0)  # Oczekiwanie na klawisz
                cv2.destroyAllWindows()

            # print(f"tabulistbefore = {tabu_list}")
            # if tabu_tenure >= 50:
            # tabu_tenure = max(15, tabu_tenure - 10)
            # print(f"tabulistafter = {tabu_list}")
            # Dodanie ruchu do listy tabu
            tabu_list.append(best_move_edges)
            if len(tabu_list) > tabu_tenure:
                tabu_list.pop(0)  # Ograniczenie dlugości listy tabu
                tabu_list.pop(0)
            # print(f"tabulistafter22 = {tabu_list}")
            history.add(tuple(current_solution))
            #save_to_excel(iteration+1, time.time() - start_time, current_cost)

            print(
                f"Iteration {iteration + 1}: Best Cost = {global_best_cost}, Current = {current_cost}, Tabu Tenure = {tabu_tenure}, Max No Improve = {max_no_improve}")

            if no_improve_count > max_no_improve:
                #print(f"Iteration {iteration + 1}: No improvement for {max_no_improve} steps, applying diversification.")
                current_solution = diversify_solution(current_solution, diversification_factor)
                current_cost = calculate_cost(G, current_solution)
                no_improve_count = 0
                best_cost = current_cost

            if current_cost < global_best_cost:
                global_best_cost = current_cost
                global_best_solution = current_solution


        iteration += 1


        # if global_best_cost <= 7544.36590190409 * 1.05 or iteration == max_iterations:
        #     save_to_excel(iteration, time.time() - start_time, global_best_cost)
        #     break
        # if global_best_cost <= 1610.0 or iteration == max_iterations:
        #     save_to_excel(iteration, time.time() - start_time, global_best_cost)
        #     break

        if keyboard.is_pressed('q'):
            break

    #print("Tabu Search completed.")

    return global_best_solution, global_best_cost


# Rysowanie grafu z aktualnym rozwiązaniem
def draw(Graph, current_solution, image):
    # Używamy spring_layout do rozplanowania wierzchołków
    pos = nx.spring_layout(Graph, seed=42, scale=0.8)
    width, height = 1600, 800  # Ustalamy wymiary obrazu
    node_positions = {k: (int(v[0] * width / 2 + width / 2), int(v[1] * height / 2 + height / 2)) for k, v in
                      pos.items()}

    # Rysowanie krawędzi
    image.fill(0)  # Czyszczenie obrazu na każdą iterację
    # for edge in Graph.edges():
    #     start_node = edge[0]
    #     end_node = edge[1]
    #     start_pos = node_positions[start_node]
    #     end_pos = node_positions[end_node]
    #     cv2.line(image, start_pos, end_pos, (192, 192, 192), 2)

    # Rysowanie wierzchołków
    for node, (x, y) in node_positions.items():
        cv2.circle(image, (x, y), 20, (135, 206, 235), -1)  # Kolor wierzchołków
        cv2.putText(image, str(node), (x - 10, y + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

    # Rysowanie trasy
    for i in range(len(current_solution) - 1):
        start_pos = node_positions[current_solution[i]]
        end_pos = node_positions[current_solution[i + 1]]
        cv2.line(image, start_pos, end_pos, (255, 0, 0), 4)  # Kolor trasy (czerwony)

    # Dodać linię powrotną do pierwszego wierzchołka
    start_pos = node_positions[current_solution[-1]]
    end_pos = node_positions[current_solution[0]]
    cv2.line(image, start_pos, end_pos, (255, 0, 0), 4)

    # Wyświetlanie grafu w oknie
    cv2.imshow('Graph Visualization', image)
    cv2.waitKey(0)  # Zamiast 200 ms, teraz tylko 1 ms, aby odświeżyć okno
    # Nie używamy cv2.destroyAllWindows() - żeby nie zamykać okna po każdej iteracji
    # Bez cv2.destroyAllWindows() - żeby nie zamykać okna po każdej iteracji


def load(path):
    tree = ET.parse(path)
    root = tree.getroot()
    G = nx.DiGraph()
    vertex_id = 0

    for vertex in root.findall('.//graph//vertex'):
        for edge in vertex.findall('edge'):
            edge_cost = float(edge.attrib.get('cost', 0))
            to_vertex = edge.text.strip()
            G.add_edge(vertex_id, int(to_vertex), weight=edge_cost)
        vertex_id += 1

    return G


import openpyxl
from openpyxl import Workbook


def save_to_excel(iteration, time_elapsed, best_cost):
    # Sprawdzamy, czy plik istnieje
    try:
        wb = load_workbook('results5.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        # Ustawianie nagłówków, jeśli plik nie istnieje
        sheet.append(['Iteration', 'Time Elapsed (s)', 'Best Cost'])

    # Szukamy pierwszego pustego wiersza, aby dodać nowe dane
    first_empty_row = sheet.max_row + 1  # Pierwszy wolny wiersz

    # Zapisujemy dane w trzech kolumnach w pierwszym wolnym wierszu
    sheet.cell(row=first_empty_row, column=1).value = iteration
    sheet.cell(row=first_empty_row, column=2).value = time_elapsed
    sheet.cell(row=first_empty_row, column=3).value = best_cost

    # Zapisz plik
    wb.save('results5.xlsx')


def generate_graph(num_nodes, img_size=500):
    G = nx.Graph()

    # Losowe współrzędne wierzchołków w zakresie (50, img_size - 50)
    positions = {i: (np.random.randint(50, img_size - 50), np.random.randint(50, img_size - 50)) for i in
                 range(num_nodes)}

    # Dodanie wierzchołków
    for node, pos in positions.items():
        G.add_node(node, pos=pos)

    # Dodanie krawędzi z wagami
    for i in range(num_nodes):
        for j in range(i + 1, num_nodes):
            dist = np.linalg.norm(np.array(positions[i]) - np.array(positions[j]))
            G.add_edge(i, j, weight=dist)

    return G, positions


# def draw_graph_cv2(G, positions, current_solution=None, img_size=600):
#     img = np.ones((img_size, img_size, 3), dtype=np.uint8) * 255
#
#
#     for (i, j) in G.edges():
#         cv2.line(img, positions[i], positions[j], (200, 200, 200), 2)  # Szare
#     # trasa
#     if current_solution:
#         for i in range(len(current_solution) - 1):
#             start_pos = positions[current_solution[i]]
#             end_pos = positions[current_solution[i + 1]]
#             cv2.line(img, start_pos, end_pos, (0, 0, 255), 3)  # Czerwona
#
#         # ostatni  z pierwszym
#         start_pos = positions[current_solution[-1]]
#         end_pos = positions[current_solution[0]]
#         cv2.line(img, start_pos, end_pos, (0, 0, 255), 3)  # Czerwona
#
#     # rysowanie w
#     for node, pos in positions.items():
#         cv2.circle(img, pos, 10, (0, 255, 0), -1)  # Zielone
#         cv2.putText(img, str(node), (pos[0] + 10, pos[1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
#
#     # wyswietlenie
#     cv2.imshow("Graph", img)
#     cv2.waitKey(1)
#

def draw_graph_cv2(G, positions, current_solution=None, img_size=550, map_image_path="mapa_polski.png"):
    # Załaduj mapę jako tło
    map_img = cv2.imread(map_image_path)
    map_img = cv2.resize(map_img, (img_size, img_size))  # Dopasuj rozmiar mapy do obrazu

    # Jeśli obraz mapy nie istnieje, stwórz białe tło
    if map_img is None:
        map_img = np.ones((img_size, img_size, 3), dtype=np.uint8) * 255

    # Nanieś krawędzie grafu
    # for (i, j) in G.edges():
        # cv2.line(map_img, positions[i], positions[j], (200, 200, 200), 2)  # Szare

    # Rysowanie trasy (jeśli podano)
    if current_solution:
        for i in range(len(current_solution) - 1):
            start_pos = positions[current_solution[i]]
            end_pos = positions[current_solution[i + 1]]
            cv2.line(map_img, start_pos, end_pos, (0, 0, 255), 3)  # Czerwona

        # Ostatni z pierwszym
        start_pos = positions[current_solution[-1]]
        end_pos = positions[current_solution[0]]
        cv2.line(map_img, start_pos, end_pos, (0, 0, 255), 3)  # Czerwona

    # Rysowanie węzłów grafu
    for node, pos in positions.items():
        cv2.circle(map_img, pos, 10, (0, 255, 0), -1)  # Zielone
        cv2.putText(map_img, str(node), (pos[0] + 10, pos[1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)

    # Wyświetlenie obrazu z mapą i grafem
    cv2.imshow("Graph on Map", map_img)
    cv2.waitKey(1)



if __name__ == "__main__":
    path = "E:/TSP_TABU/pythonProject1/TSP-TabuSearch/bayg29.xml"
    G = load(path)
    image = np.zeros((800, 1600, 3), dtype=np.uint8)
    try:
        num_nodes = 50  # Liczba wierzchołków
        G, positions = generate_graph(num_nodes)
        draw_graph_cv2(G, positions)
        cv2.waitKey(0)  # Oczekiwanie na klawisz
        cv2.destroyAllWindows()
        start_time = time.time()
        best_solution, best_cost = tabu_search(G, max_iterations=1000, diversification_factor=0.15)
        draw_graph_cv2(G, positions, best_solution)
        cv2.waitKey(0)  # Oczekiwanie na klawisz
        cv2.destroyAllWindows()
       # for i in range(201):
       #  print(i)
       #  start_time = time.time()
       #  best_solution, best_cost = tabu_search(G, max_iterations=1000, diversification_factor=0.15)
       #  end_time = time.time()
       #  execution_time = end_time - start_time
       #  #print(f"Time taken for Tabu Search: {execution_time:.2f} seconds")
       #  #print("Best Solution (Tour):", best_solution)
       #  #print("Best Cost:", best_cost)
       #  #draw(G, best_solution, image)
    except ValueError as e:
        print(e)
